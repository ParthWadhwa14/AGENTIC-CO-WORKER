from pathlib import Path
from uuid import uuid4

import pandas as pd
from openpyxl import load_workbook

from app.models import ParsedElement


class XLSXLoader:
    source_type = "xlsx"

    def load(
        self,
        file_path: str,
        document_id: str | None = None,
        rows_per_chunk: int = 25
    ) -> list[ParsedElement]:
        document_id = document_id or str(uuid4())
        path = Path(file_path)

        workbook = load_workbook(file_path, data_only=False)
        elements: list[ParsedElement] = []

        # Workbook summary
        sheet_names = workbook.sheetnames
        workbook_summary = (
            f"Excel workbook: {path.name}\n"
            f"Sheets: {', '.join(sheet_names)}"
        )

        elements.append(
            ParsedElement(
                document_id=document_id,
                source_type=self.source_type,
                file_name=path.name,
                element_type="workbook_summary",
                text=workbook_summary,
                metadata={
                    "file_path": str(path),
                    "sheet_names": sheet_names,
                },
            )
        )

        # Process each sheet using pandas
        all_sheets = pd.read_excel(file_path, sheet_name=None)

        for sheet_name, df in all_sheets.items():
            df = df.fillna("")

            if df.empty:
                continue

            schema_text = self._build_sheet_schema_text(path.name, sheet_name, df)

            elements.append(
                ParsedElement(
                    document_id=document_id,
                    source_type=self.source_type,
                    file_name=path.name,
                    sheet_name=sheet_name,
                    element_type="sheet_schema",
                    text=schema_text,
                    metadata={
                        "file_path": str(path),
                        "columns": list(df.columns),
                        "row_count": len(df),
                    },
                )
            )

            for start in range(0, len(df), rows_per_chunk):
                end = min(start + rows_per_chunk, len(df))
                sub_df = df.iloc[start:end]

                row_text = self._rows_to_text(sub_df, start_index=start)

                elements.append(
                    ParsedElement(
                        document_id=document_id,
                        source_type=self.source_type,
                        file_name=path.name,
                        sheet_name=sheet_name,
                        element_type="row_group",
                        text=row_text,
                        row_start=start + 1,
                        row_end=end,
                        metadata={
                            "file_path": str(path),
                            "columns": list(df.columns),
                            "row_count": len(df),
                        },
                    )
                )

            # Formula chunks
            ws = workbook[sheet_name]
            formula_texts = []

            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_texts.append(
                            f"Cell {cell.coordinate}: {cell.value}"
                        )

            if formula_texts:
                elements.append(
                    ParsedElement(
                        document_id=document_id,
                        source_type=self.source_type,
                        file_name=path.name,
                        sheet_name=sheet_name,
                        element_type="formulas",
                        text="\n".join(formula_texts),
                        metadata={
                            "file_path": str(path),
                            "formula_count": len(formula_texts),
                        },
                    )
                )

        return elements

    def _build_sheet_schema_text(
        self,
        file_name: str,
        sheet_name: str,
        df: pd.DataFrame
    ) -> str:
        lines = [
            f"Excel workbook: {file_name}",
            f"Sheet: {sheet_name}",
            f"Rows: {len(df)}",
            "Columns:",
        ]

        for col in df.columns:
            dtype = str(df[col].dtype)
            sample_values = df[col].dropna().astype(str).head(5).tolist()
            lines.append(
                f"- {col} | type: {dtype} | sample values: {sample_values}"
            )

        return "\n".join(lines)

    def _rows_to_text(self, df: pd.DataFrame, start_index: int) -> str:
        lines = []

        for local_i, (_, row) in enumerate(df.iterrows()):
            row_number = start_index + local_i + 1
            row_parts = [f"{col}: {row[col]}" for col in df.columns]
            lines.append(f"Row {row_number}: " + " | ".join(row_parts))

        return "\n".join(lines)